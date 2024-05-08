from openpyxl import load_workbook

path=("ActiveEmployeeData.xlsx")

wb=load_workbook(path)

sheet=wb.active

row=sheet.max_row
col=sheet.max_column

for i in range(1,row+1):
    for j in range(1,col+1):
        cell_obj=sheet.cell(row=i,col=j)
        print(cell_obj.value)

with open('work/payroll.payroll.txt','r') as r,open('account/payroll.payroll.txt') as w:
    w.write(r.read())

payroll_data = {}
 
with open(payroll_file_path, 'r') as file:
        
    for line in file:

        emp_code, work_days = line.strip().split(',') # print("Employee code:", emp_code,"|| Work days :", work_days)
        salary = int(work_days) * 1000
        payroll_data[emp_code] = salary

    # Read employee data and create a workbook
    wb = load_workbook(employee_data_path)
    sh1 = wb['Sheet1']
    rows = sh1.max_row
    columns = sh1.max_column 

    for i, value in enumerate(payroll_data.values(), start=2):
        cell = sh1.cell(row=i, column=8)
        cell.value = value

    wb.save(employee_data_path)

print(row)
print(col)

x1=sheet.cell(row=1,column=8)
x1.value="salary"
