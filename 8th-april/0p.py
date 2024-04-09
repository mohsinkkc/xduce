from openpyxl import load_workbook
import shutil
import os

# from openpyxl import  workbook
path="ActiveEmployeeData.xlsx"

wb_obj= load_workbook(path)
# workbook=workbook()
# workbook.save(filename="ActiveEmployeeData.xlsx")
# wb=openpyxl.workbook()
# sheet=wb.active


sheet_obj=wb_obj.active

row_count = sheet_obj.max_row
col_count = sheet_obj.max_column

# breakpoint()

for i in range(1,row_count+1):
    for j in range(1,col_count+1):
        cell_obj=sheet_obj.cell(row=i,column=j)
        print(cell_obj.value)



with open('account/payroll.txt','r') as r, open('work/payroll.txt','w') as w:
    w.write(r.read())

# x1=sheet_obj.cell(row=1,column=8)
# x1.value = "salary"

salary_column = col_count + 1  
salary_header = "Salary"  
sheet_obj.cell(row=1, column=salary_column, value=salary_header) 

with open('account/payroll.txt', 'r') as f:
    salaries = f.readlines()  
    


payroll_data = {}
 
with open("account\payroll.txt", 'r') as file:
        
    for line in file:

        emp_code, work_days = line.strip().split(',')
        salary = int(work_days) * 1000
        payroll_data[emp_code] = salary

    wb = load_workbook("ActiveEmployeeData.xlsx")
    sh1 = wb['Sheet1']
    rows = sh1.max_row
    columns = sh1.max_column 

    with open('C:/Users/Mohasin Mohmad/Desktop/mohsin python/8th-april/account/payroll.txt', 'r') as file:
        lines = file.readlines()


    for i, line in enumerate(lines, start=1):
        print(f"Line {i}: {line.strip()}") 

    wb.save("ActiveEmployeeData.xlsx")
print(row_count)
print(col_count)
    
wb_obj.save("ActiveEmployeeData.xlsx")


# def copy_excel_to_folder(source_file, destination_folder):
#     if not os.path.isfile(source_file):
#         print(f"Error: Source file '{source_file}' does not exist.")
#         return False

#     if not os.path.isdir(destination_folder):
#         print(f"Error: Destination folder '{destination_folder}' does not exist.")
#         return False

#     try:
        
#         shutil.copy(source_file, destination_folder)
#         print(f"File '{os.path.basename(source_file)}' copied successfully to '{destination_folder}'.")
#         return True
#     except Exception as e:
#         print(f"Error: Failed to copy file - {e}")
#         return False

# source_file = 'C:/Users/Mohasin Mohmad/Desktop/mohsin python/8th-april/payroll.payroll.txt'
# destination_folder = 'C:/Users/Mohasin Mohmad/Desktop/mohsin python/8th-april/Destination'
# copy_excel_to_folder(source_file, destination_folder)


